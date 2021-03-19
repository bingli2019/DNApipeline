#! /usr/bin/python3
# -*- coding: utf-8 -*-

__Version__ = "0.18"
__Author__ = "pzweuj"
__Date__ = "20210310"


import os
import time
import sys
import yaml
from openpyxl import Workbook
from docx2pdf import convert

# 配置文件读入
def getRunningInfo(runInfo):
    with open(runInfo, "r") as stream:
        try:
            configDict = yaml.safe_load(stream)
            return configDict
        except yaml.YAMLError as exc:
            print(exc)

# 获得当前时间
def getTime():
    localTime = time.asctime(time.localtime(time.time()))
    return localTime

# 获得脚本路径
def getAbsPath():
    now = os.path.abspath(os.path.dirname(sys.argv[0]))
    return now

# 新建文件夹
def mkdir(folder_path):
    folder = os.path.exists(folder_path)
    if not folder:
        os.makedirs(folder_path)
        print("Create new folder " + folder_path + " done!")
    else:
        print("Folder " + folder_path + " already exists!")

# 获得当前路径
def getCurrentPath():
    return os.getcwd()

# 聚合结果
def read_txt(txt):
    txtFile = open(txt, "r", encoding="utf-8")
    l = []
    for i in txtFile:
        l.append(i.split("\n")[0])
    txtFile.close()
    return l

def write_excel(ws, contents):
    for i in range(len(contents)):
        col = contents[i].split("\t")
        for c in range(len(col)):
            cstring = col[c]
            cell = ws.cell(row=1+i, column=1+c)
            cell.value = cstring

def mergeResultsToExcel(resultsDir, sampleID):
    dir_list = os.listdir(resultsDir)
    wb = Workbook()

    if "QC" in dir_list:
        qc = os.listdir(resultsDir + "/QC")
        if len(qc) != 0:
            for q in qc:
                content = read_txt(resultsDir + "/QC/" + q)
                ws_qc = wb.create_sheet(q.replace(".txt", "").replace(sampleID + ".", ""))
                write_excel(ws_qc, content)

    if "annotation" in dir_list:
        anno = os.listdir(resultsDir + "/annotation")
        if len(anno) != 0:
            for a in anno:
                content = read_txt(resultsDir + "/annotation/" + a)
                ws_anno = wb.create_sheet(a.replace(".txt", "").replace(sampleID + ".", ""))
                write_excel(ws_anno, content)

    if "cnv" in dir_list:
        cnv = os.listdir(resultsDir + "/cnv")
        if len(cnv) != 0:
            for c in cnv:
                content = read_txt(resultsDir + "/cnv/" + c)
                ws_cnv = wb.create_sheet(c.replace(".txt", "").replace(sampleID + ".", ""))
                write_excel(ws_cnv, content)

    if "sv" in dir_list:
        sv = os.listdir(resultsDir + "/sv")
        if len(sv) != 0:
            for s in sv:
                if ".txt" in s:
                    content = read_txt(resultsDir + "/sv/" + s)
                    ws_sv = wb.create_sheet(s.replace(".txt", "").replace(sampleID + ".", ""))
                    write_excel(ws_sv, content)

    if "msi" in dir_list:
        msi = os.listdir(resultsDir + "/msi")
        if len(msi) != 0:
            for m in msi:
                content = read_txt(resultsDir + "/msi/" + m)
                ws_msi = wb.create_sheet(m.replace(".txt", "").replace(sampleID + ".", ""))
                write_excel(ws_msi, content)

    wb.remove(wb["Sheet"])
    wb.save(resultsDir + "/" + sampleID + ".xlsx")
    print(sampleID + " 结果已汇总到excel表格中： " + resultsDir + "/" + sampleID + ".xlsx")


def convertDOCX2PDF(input, output):
    convert(input, output)